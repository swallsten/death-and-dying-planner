"""Death Planning Workbook — a Streamlit app.

Run: streamlit run app.py
"""
from __future__ import annotations

import base64
import io
import json
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import streamlit as st
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)

APP_DIR = Path(__file__).parent
XLSX_PATH = APP_DIR / "Death_Planning_Workbook.xlsx"
DATA_DIR = APP_DIR / "data"
ATTACH_DIR = DATA_DIR / "attachments"

# Detect cloud deployment — Streamlit Community Cloud sets specific env vars.
# STREAMLIT_SERVER_HEADLESS is also set locally by `streamlit run`, so it's not
# reliable on its own. We check for vars that only exist on Community Cloud.
IS_CLOUD = bool(
    os.environ.get("STREAMLIT_SHARING_MODE")
    or os.environ.get("HOSTNAME", "").endswith(".streamlit.app")
    or os.environ.get("HOME", "").startswith("/mount/")
    or os.environ.get("STREAMLIT_SERVER_ADDRESS")  # set on Community Cloud
)

# Only create disk dirs when running locally
if not IS_CLOUD:
    DATA_DIR.mkdir(exist_ok=True)
    ATTACH_DIR.mkdir(exist_ok=True)

CHECKLIST_SHEETS = ["Before", "During Illness", "At End - Hospice", "After Death"]
TABLE_SHEETS = ["Key Contacts", "Account Registry"]
STATUSES = ["Not started", "In progress", "Done", "N/A"]
KDF_ITERATIONS = 200_000

# ---------------- Forms registry ----------------
# Fillable templates. HIPAA is federal and actually standardizable.
# The others are personal (non-legal) documents.
# State-specific forms (advance directive, POA, DNR) are NOT generated —
# we link out to the state's official version instead.

FORMS = {
    "hipaa": {
        "title": "HIPAA Authorization for Release of Health Information",
        "kind": "legal-federal",
        "blurb": (
            "A standard authorization under 45 CFR 164.508. When signed, it lets a "
            "healthcare provider release your medical records to a person you name. "
            "Valid nationwide."
        ),
        "fields": [
            ("patient_name", "Patient full name", "text", ""),
            ("patient_dob", "Patient date of birth", "text", ""),
            ("patient_address", "Patient address", "textarea", ""),
            ("patient_phone", "Patient phone", "text", ""),
            ("provider_name", "Provider / facility releasing records", "textarea",
             "Name and address of the hospital, clinic, doctor, or health plan that currently holds the records."),
            ("recipient_name", "Person / entity receiving the records", "textarea",
             "Full name, relationship, and address of the person authorized to receive the records."),
            ("info_description", "Information to be released", "textarea",
             "All medical records from [date] to [date], including but not limited to "
             "history and physicals, progress notes, lab and imaging results, medication lists, "
             "discharge summaries, and billing records."),
            ("purpose", "Purpose of disclosure", "textarea",
             "At the request of the individual, for personal use and care coordination."),
            ("expiration", "Expiration date or event", "text",
             "One year from the date of signature, or upon written revocation."),
            ("inc_psych", "Include mental health / psychotherapy notes", "checkbox", False),
            ("inc_hiv", "Include HIV / AIDS-related information", "checkbox", False),
            ("inc_substance", "Include alcohol / substance use treatment records", "checkbox", False),
            ("inc_genetic", "Include genetic information", "checkbox", False),
            ("signer_is_representative", "Signed by a personal representative (not the patient)", "checkbox", False),
            ("representative_name", "Personal representative name (if applicable)", "text", ""),
            ("representative_relationship", "Relationship / legal authority", "text", ""),
        ],
    },
    "letter_of_instruction": {
        "title": "Letter of Instruction",
        "kind": "personal",
        "blurb": (
            "A plain-English letter to your family and executor. Not a legal document — "
            "it complements the will. This is where you say, in your own words, where "
            "everything is and who to call."
        ),
        "fields": [
            ("writer_name", "Your name", "text", ""),
            ("date", "Date", "text", ""),
            ("addressed_to", "Addressed to", "text", "To my family and executor"),
            ("opening", "Opening message", "textarea",
             "If you are reading this, something has happened to me. I'm sorry I can't be "
             "there to walk you through this. Here is what you need to know."),
            ("location_of_will", "Where the will and trust documents are", "textarea", ""),
            ("location_of_financial", "Where financial records are", "textarea", ""),
            ("location_of_passwords", "Where passwords / password manager master info is", "textarea",
             "Do not put the master password here. Describe where the trusted person can find it."),
            ("first_calls", "Who to call first", "textarea",
             "Estate attorney, financial advisor, close family. Names and phone numbers."),
            ("financial_summary", "Summary of accounts and assets", "textarea",
             "Major accounts, real estate, retirement, insurance. Keep the detail in the Account Registry."),
            ("debts", "Debts and ongoing bills", "textarea", ""),
            ("digital_assets", "Digital accounts and instructions", "textarea",
             "Email, social media, subscriptions. What to do with each."),
            ("funeral_wishes", "Funeral and burial wishes (summary)", "textarea", ""),
            ("personal_messages", "Personal messages to loved ones", "textarea", ""),
            ("closing", "Closing", "textarea",
             "Thank you for handling this. I love you."),
        ],
    },
    "funeral_wishes": {
        "title": "Funeral and Burial Wishes",
        "kind": "personal",
        "blurb": (
            "A short statement of preferences. Not legally binding, but hugely helpful "
            "for the people making decisions in the first 48 hours."
        ),
        "fields": [
            ("name", "Your name", "text", ""),
            ("date", "Date", "text", ""),
            ("disposition", "Burial, cremation, donation, or other", "text", ""),
            ("location", "Cemetery, scattering location, or facility", "textarea", ""),
            ("prepaid", "Pre-paid arrangements, if any (funeral home, plot, etc.)", "textarea", ""),
            ("service_type", "Service type (religious, secular, memorial, none)", "text", ""),
            ("officiant", "Preferred officiant or clergy", "text", ""),
            ("location_of_service", "Preferred service location", "text", ""),
            ("music", "Music preferences", "textarea", ""),
            ("readings", "Readings, speakers, or prayers", "textarea", ""),
            ("flowers_donations", "Flowers or charitable donations", "textarea", ""),
            ("obituary", "Obituary notes or draft", "textarea", ""),
            ("notify", "People or organizations to notify", "textarea", ""),
            ("additional", "Anything else", "textarea", ""),
        ],
    },
    "digital_assets": {
        "title": "Digital Asset Instructions",
        "kind": "personal",
        "blurb": (
            "Instructions for what to do with your email, social media, subscriptions, "
            "and other online accounts. Not legally binding on its own, but many "
            "platforms honor these when paired with a death certificate."
        ),
        "fields": [
            ("writer_name", "Your name", "text", ""),
            ("date", "Date", "text", ""),
            ("password_manager", "Password manager and how to access it", "textarea",
             "Do not put the master password here. Describe where it lives (e.g., sealed envelope in safe deposit box)."),
            ("primary_email", "Primary email account and instructions", "textarea",
             "What to do with the email account (preserve, close, forward)."),
            ("phone_accounts", "Phone and 2FA accounts", "textarea",
             "Carrier, phone number, SIM, and how to maintain access for 2FA during transition."),
            ("financial_logins", "Financial account logins", "textarea",
             "Reference only — actual credentials belong in the password manager."),
            ("social_media", "Social media accounts (Facebook, Instagram, LinkedIn, X, etc.)", "textarea",
             "Memorialize, delete, or leave active. Instructions per account."),
            ("cloud_storage", "Cloud storage and photos (iCloud, Google Photos, Dropbox)", "textarea", ""),
            ("subscriptions", "Subscriptions to cancel", "textarea",
             "Reference the Account Registry for the full list."),
            ("crypto", "Cryptocurrency, wallets, keys", "textarea",
             "Where seed phrases / hardware wallets live. Never put the phrase itself here."),
            ("legacy_contacts", "Legacy contacts already configured", "textarea",
             "Apple Legacy Contact, Google Inactive Account Manager, Facebook Legacy Contact, etc."),
            ("notes", "Other notes", "textarea", ""),
        ],
    },
}

STATE_FORM_POINTERS = [
    ("Advance Directive / Living Will",
     "State-specific. Download your state's official form.",
     "Search: [your state] advance directive form. AARP and CaringInfo maintain free state-by-state lists."),
    ("Healthcare Power of Attorney",
     "State-specific. Often combined with the advance directive.",
     "Search: [your state] healthcare power of attorney form."),
    ("Durable Financial Power of Attorney",
     "State-specific. Usually requires notarization; some states require witnesses.",
     "Work with an estate attorney or use your state bar association's form."),
    ("DNR / POLST / MOLST",
     "State-specific. Must be signed by a physician. Cannot be generated generically.",
     "Ask your primary care physician or hospice team. Forms vary by state name (POLST, MOLST, POST, etc.)."),
]

# Modes
MODES = ["For my parents", "For ourselves", "For someone else"]

# Muted blue/gray palette
PRIMARY = "#4A6274"
MUTED = "#7A8B99"
BG = "#F5F7FA"


# ---------------- Mode context ----------------
# All UI text that should shift based on who this workbook is for.
# "subject" = the person whose affairs are being organized.
# "user" = the person filling out the workbook.

def _someone_name() -> str:
    return st.session_state.get("someone_name", "this person") or "this person"


def ctx() -> dict:
    """Return a context dict for the current mode — used for all UI text."""
    mode = mode_key()
    if "parent" in mode.lower():
        return {
            "subject": "your parent",
            "subject_plural": "your parents",
            "subject_possessive": "their",
            "subject_they": "they",
            "user_role": "adult child",
            "you_or_they": "they",
            "your_or_their": "their",
            "you_or_them": "them",
            "ask_or_write": "Ask about",
            "gather_verb": "Find out where they keep",
            "before_urgency": (
                "This is the most critical section. Everything here requires your parent "
                "to be alive and mentally competent. Once that changes, most of these "
                "become impossible or extremely difficult. Have the conversation now."
            ),
            "during_intro": (
                "Your parent is declining and you're managing their care. This section "
                "helps you track the logistical reality of caregiving."
            ),
            "end_intro": (
                "Your parent is near the end. These are the decisions and arrangements "
                "that need to happen now."
            ),
            "after_intro": (
                "Your parent has died. Here's what needs to happen and roughly when. "
                "The first 48 hours are the most time-sensitive."
            ),
            "contacts_intro": (
                "The professionals and institutions involved in your parent's affairs. "
                "You'll be calling all of these people."
            ),
            "registry_intro": (
                "Every account, policy, and subscription your parent has. "
                "You'll need this to close, transfer, or cancel each one."
            ),
            "forms_hipaa": (
                "For this mode, **you** are the recipient — the person getting access to "
                "your parent's medical records. Your parent (or their legal representative) signs it."
            ),
            "forms_letter": (
                "This is something your parent would write, not you. If they're willing, "
                "sit with them and help them fill it out. If not, fill in what you know "
                "and note what's missing."
            ),
            "forms_funeral": (
                "Ask your parent about their preferences. Writing it down now saves you from "
                "guessing during the worst 48 hours of your life."
            ),
            "forms_digital": (
                "Where are their accounts? How do you get in? What do they want done "
                "with their email, social media, photos? Ask now."
            ),
            "report_before_flag": (
                "These items require a conversation with your parent while they're able to participate. "
                "Every one left open is a gap you'll have to fill blind."
            ),
            "overview_intro": (
                "You're organizing your parent's affairs — gathering the information "
                "you'll need when the time comes, or that you need right now if things are "
                "already in motion."
            ),
        }
    elif "ourselves" in mode.lower():
        return {
            "subject": "you",
            "subject_plural": "you and your partner",
            "subject_possessive": "your",
            "subject_they": "you",
            "user_role": "the person preparing",
            "you_or_they": "you",
            "your_or_their": "your",
            "you_or_them": "you",
            "ask_or_write": "Document",
            "gather_verb": "Write down where you keep",
            "before_urgency": (
                "This is the gift you give your kids (or whoever handles your affairs). "
                "Everything here is something *you* know right now but nobody else does. "
                "If you get hit by a bus tomorrow, can they find it all? Do this section first."
            ),
            "during_intro": (
                "If you're managing your own illness or decline, these are the systems "
                "and people you'll put in place. If you're preparing in advance, think of "
                "this as instructions to whoever steps in."
            ),
            "end_intro": (
                "Your wishes for the very end. Writing this down is an act of kindness — "
                "it takes impossible decisions off your family's plate."
            ),
            "after_intro": (
                "What your family will need to do after you die, and roughly when. "
                "Reviewing this now helps you make sure the information they'll need "
                "is actually documented."
            ),
            "contacts_intro": (
                "The people your family will need to call. Write them down here so "
                "your kids don't have to go hunting."
            ),
            "registry_intro": (
                "Every account, policy, and subscription you have. Your family will "
                "need to close, transfer, or cancel each one. They won't know what "
                "exists unless you tell them."
            ),
            "forms_hipaa": (
                "For this mode, **your kids or designated person** are the recipient — "
                "you're granting them access to your medical records. You sign it."
            ),
            "forms_letter": (
                "This is the core deliverable of this workbook — a letter from you to "
                "your family, in your own words, explaining where everything is and "
                "who to call. Take your time with it."
            ),
            "forms_funeral": (
                "Write down what you want. Burial or cremation? What kind of service? "
                "Your family will be grateful they don't have to guess."
            ),
            "forms_digital": (
                "Document your digital life. Where's the password manager? What happens to "
                "your email, social media, photos? Your family won't know any of this."
            ),
            "report_before_flag": (
                "These are things only you know. Every one left open is something your "
                "family will have to figure out without you."
            ),
            "overview_intro": (
                "You're getting your own affairs in order — creating the binder your "
                "kids or executor will need when the time comes. This is one of the "
                "most useful things you can do for the people you love."
            ),
        }
    else:
        name = _someone_name()
        return {
            "subject": name,
            "subject_plural": name,
            "subject_possessive": f"{name}'s",
            "subject_they": "they",
            "user_role": "the person helping",
            "you_or_they": "they",
            "your_or_their": "their",
            "you_or_them": "them",
            "ask_or_write": "Find out",
            "gather_verb": f"Find out where {name} keeps",
            "before_urgency": (
                f"This is the most critical section. Everything here requires {name} "
                "to be alive and mentally competent. Once that changes, most of these "
                "become impossible or extremely difficult."
            ),
            "during_intro": (
                f"{name} is declining and you're helping manage their care. This section "
                "helps you track the logistics."
            ),
            "end_intro": (
                f"{name} is near the end. These are the decisions and arrangements "
                "that need to happen now."
            ),
            "after_intro": (
                f"{name} has died. Here's what needs to happen and roughly when."
            ),
            "contacts_intro": (
                f"The professionals and institutions involved in {name}'s affairs."
            ),
            "registry_intro": (
                f"Every account, policy, and subscription {name} has."
            ),
            "forms_hipaa": (
                f"**You** (or the designated person) are the recipient — getting access to "
                f"{name}'s medical records. {name} (or their legal representative) signs it."
            ),
            "forms_letter": (
                f"This is something {name} would write. If they're able, help them fill it "
                "out. Otherwise, fill in what you know and note the gaps."
            ),
            "forms_funeral": (
                f"Ask {name} about their preferences, if possible. Otherwise, record what "
                "you know or what the family has discussed."
            ),
            "forms_digital": (
                f"Where are {name}'s accounts? How do you get in? "
                "What do they want done with their digital life?"
            ),
            "report_before_flag": (
                f"These items require {name}'s participation while they're able. "
                "Every one left open is a gap."
            ),
            "overview_intro": (
                f"You're organizing {name}'s affairs — gathering the information "
                "that will be needed when the time comes."
            ),
        }


# ---------------- Parsing ----------------

@st.cache_data
def load_workbook_structure() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    out: dict = {"checklists": {}, "tables": {}, "instructions": []}

    if "Instructions" in wb.sheetnames:
        for row in wb["Instructions"].iter_rows(values_only=True):
            if row and row[0]:
                out["instructions"].append(str(row[0]))

    for name in CHECKLIST_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        items = []
        current_category = None
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            cells = [None if c is None else str(c).strip() for c in row]
            non_empty = [c for c in cells if c]
            if len(non_empty) == 1 and cells[0] and cells[0] == cells[0].upper():
                current_category = cells[0].title()
                continue

            if name == "After Death":
                timeframe = cells[0] or ""
                sub = cells[1] or ""
                task = cells[2] or ""
                details = cells[3] or ""
                if not task:
                    continue
                items.append({
                    "category": current_category or timeframe or "",
                    "timeframe": timeframe,
                    "subcategory": sub,
                    "item": task,
                    "details": details,
                    "notes": "",
                })
            else:
                cat = cells[0] or current_category or ""
                item = cells[1] or ""
                details = cells[2] or ""
                notes = cells[3] or ""
                if not item:
                    continue
                items.append({
                    "category": current_category or cat,
                    "subcategory": cat,
                    "item": item,
                    "details": details,
                    "notes": notes,
                })
        out["checklists"][name] = items

    for name in TABLE_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c else "" for c in rows[0]]
        seed = []
        current_section = None
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            cells = [None if c is None else str(c).strip() for c in row]
            non_empty = [c for c in cells if c]
            if len(non_empty) == 1 and cells[0] and cells[0] == cells[0].upper():
                current_section = cells[0].title()
                continue
            seed_row = {h: (cells[i] if i < len(cells) and cells[i] else "")
                        for i, h in enumerate(headers)}
            seed_row["_section"] = current_section or ""
            seed.append(seed_row)
        out["tables"][name] = {"headers": headers, "seed": seed}

    return out


# ---------------- Crypto ----------------

def derive_key(passphrase: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(), length=32, salt=salt, iterations=KDF_ITERATIONS,
    )
    return base64.urlsafe_b64encode(kdf.derive(passphrase.encode("utf-8")))


def get_fernet(mode: str) -> Optional[Fernet]:
    return st.session_state.get("fernets", {}).get(mode)


def set_fernet(mode: str, f: Optional[Fernet]):
    st.session_state.setdefault("fernets", {})[mode] = f


def get_salt(mode: str) -> Optional[str]:
    return st.session_state.get("salts", {}).get(mode)


def set_salt(mode: str, salt_b64: Optional[str]):
    st.session_state.setdefault("salts", {})[mode] = salt_b64


# ---------------- Persistence ----------------

def mode_key() -> str:
    return st.session_state.get("mode", "For my parents")


def mode_slug(mode: str) -> str:
    if "parent" in mode.lower():
        return "parents"
    elif "ourselves" in mode.lower():
        return "self"
    else:
        return "other"


def data_file(mode: str) -> Path:
    return DATA_DIR / f"workbook_{mode_slug(mode)}.json"


def default_user_state(structure: dict) -> dict:
    state = {"checklists": {}, "tables": {}, "forms": {}}
    for sheet, items in structure["checklists"].items():
        state["checklists"][sheet] = [
            {"status": "Not started", "user_notes": "", "user_info": "", "attachments": []}
            for _ in items
        ]
    for sheet, tbl in structure["tables"].items():
        state["tables"][sheet] = [dict(r) for r in tbl["seed"]]
    for form_id, spec in FORMS.items():
        state["forms"][form_id] = {
            f[0]: f[3] for f in spec["fields"]
        }
    return state


def reconcile(data: dict, structure: dict) -> dict:
    base = default_user_state(structure)
    for sheet, items in base["checklists"].items():
        saved = data.get("checklists", {}).get(sheet, [])
        for i, it in enumerate(items):
            if i < len(saved) and isinstance(saved[i], dict):
                for k in ("status", "user_notes", "user_info"):
                    if k in saved[i]:
                        it[k] = saved[i][k]
                if "attachments" in saved[i]:
                    it["attachments"] = saved[i]["attachments"]
    for sheet in base["tables"]:
        if sheet in data.get("tables", {}):
            base["tables"][sheet] = data["tables"][sheet]
    saved_forms = data.get("forms", {}) or {}
    for form_id in base["forms"]:
        if form_id in saved_forms and isinstance(saved_forms[form_id], dict):
            base["forms"][form_id].update({
                k: v for k, v in saved_forms[form_id].items()
                if k in base["forms"][form_id]
            })
    return base


def file_is_encrypted(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        blob = json.loads(path.read_text())
        return bool(blob.get("encrypted"))
    except Exception:
        return False


def save_user_data(mode: str):
    """Save to disk (local mode only). Returns path or None on cloud."""
    if IS_CLOUD:
        return None
    path = data_file(mode)
    data = st.session_state["user_data"][mode]
    payload = {
        "mode": mode,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "data": data,
    }
    if "someone" in mode.lower():
        payload["someone_name"] = st.session_state.get("someone_name", "")
    f = get_fernet(mode)
    if f is not None:
        salt_b64 = get_salt(mode) or base64.b64encode(os.urandom(16)).decode()
        set_salt(mode, salt_b64)
        ciphertext = f.encrypt(json.dumps(payload).encode("utf-8"))
        path.write_text(json.dumps({
            "encrypted": True,
            "kdf": "pbkdf2-sha256",
            "iterations": KDF_ITERATIONS,
            "salt": salt_b64,
            "ciphertext": ciphertext.decode("ascii"),
        }, indent=2))
    else:
        path.write_text(json.dumps(payload, indent=2))
    return path


def load_user_data(mode: str, structure: dict, passphrase: Optional[str] = None) -> tuple[dict, str]:
    """Returns (data, status): 'loaded' | 'default' | 'locked' | 'bad_passphrase'."""
    path = data_file(mode)
    if not path.exists():
        return default_user_state(structure), "default"
    try:
        blob = json.loads(path.read_text())
    except Exception as e:
        st.warning(f"Couldn't parse {path.name}: {e}")
        return default_user_state(structure), "default"

    def _restore_someone_name(payload: dict):
        if "someone" in mode.lower() and "someone_name" in payload:
            st.session_state["someone_name"] = payload["someone_name"]

    if blob.get("encrypted"):
        if not passphrase:
            return default_user_state(structure), "locked"
        try:
            salt = base64.b64decode(blob["salt"])
            key = derive_key(passphrase, salt)
            f = Fernet(key)
            plain = f.decrypt(blob["ciphertext"].encode("ascii"))
            payload = json.loads(plain)
            set_fernet(mode, f)
            set_salt(mode, blob["salt"])
            _restore_someone_name(payload)
            return reconcile(payload.get("data", {}), structure), "loaded"
        except (InvalidToken, ValueError, KeyError):
            return default_user_state(structure), "bad_passphrase"
    else:
        _restore_someone_name(blob)
        return reconcile(blob.get("data", {}), structure), "loaded"


# ---------------- Attachments ----------------
# On cloud: files live in session_state["_attach_bytes"][key] = bytes
# On local: files live on disk in data/attachments/

def _attach_mem_key(mode: str, sheet: str, stored_name: str) -> str:
    return f"{mode_slug(mode)}:{sheet}:{stored_name}"


def item_attach_dir(mode: str, sheet: str) -> Path:
    sheet_slug = sheet.lower().replace(" ", "_").replace("-", "")
    d = ATTACH_DIR / mode_slug(mode) / sheet_slug
    if not IS_CLOUD:
        d.mkdir(parents=True, exist_ok=True)
    return d


def save_attachment(mode: str, sheet: str, idx: int, uploaded) -> dict:
    uid = uuid.uuid4().hex[:10]
    safe_name = uploaded.name.replace("/", "_").replace("\\", "_")
    base_name = f"{idx:03d}_{uid}_{safe_name}"
    f = get_fernet(mode)
    data = uploaded.getvalue()
    if f is not None:
        stored_name = base_name + ".enc"
        stored_data = f.encrypt(data)
        encrypted = True
    else:
        stored_name = base_name
        stored_data = data
        encrypted = False

    if IS_CLOUD:
        st.session_state.setdefault("_attach_bytes", {})[
            _attach_mem_key(mode, sheet, stored_name)
        ] = stored_data
    else:
        d = item_attach_dir(mode, sheet)
        (d / stored_name).write_bytes(stored_data)

    return {
        "id": uid,
        "name": uploaded.name,
        "stored_name": stored_name,
        "size": len(data),
        "encrypted": encrypted,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
    }


def read_attachment(mode: str, sheet: str, att: dict) -> Optional[bytes]:
    stored_name = att["stored_name"]
    if IS_CLOUD:
        raw = st.session_state.get("_attach_bytes", {}).get(
            _attach_mem_key(mode, sheet, stored_name)
        )
        if raw is None:
            return None
    else:
        path = item_attach_dir(mode, sheet) / stored_name
        if not path.exists():
            return None
        raw = path.read_bytes()

    if att.get("encrypted"):
        f = get_fernet(mode)
        if f is None:
            return None
        try:
            return f.decrypt(raw)
        except InvalidToken:
            return None
    return raw


def delete_attachment(mode: str, sheet: str, att: dict):
    stored_name = att["stored_name"]
    if IS_CLOUD:
        st.session_state.get("_attach_bytes", {}).pop(
            _attach_mem_key(mode, sheet, stored_name), None
        )
    else:
        path = item_attach_dir(mode, sheet) / stored_name
        if path.exists():
            path.unlink()


def export_portable_json(mode: str) -> dict:
    """Build a fully self-contained JSON-serializable dict with attachment bytes embedded.

    Attachments are stored as base64 under each item's "attachments" list, in a
    key called "data_b64". The raw file bytes are always plaintext (decrypted if
    needed) so the portable JSON can be loaded anywhere.
    """
    import copy
    user = copy.deepcopy(st.session_state["user_data"][mode])
    for sheet, items in user.get("checklists", {}).items():
        for item in items:
            for att in item.get("attachments", []):
                raw = read_attachment(mode, sheet, att)
                if raw is not None:
                    att["data_b64"] = base64.b64encode(raw).decode("ascii")
                else:
                    att["data_b64"] = None
    obj: dict = {"mode": mode, "data": user}
    if "someone" in mode.lower():
        obj["someone_name"] = st.session_state.get("someone_name", "")
    return obj


def import_portable_json(blob: dict, mode: str, structure: dict):
    """Load a portable JSON (possibly with embedded attachment bytes).

    Extracts any base64 attachment data and stores it via save_attachment's
    storage backend (disk or memory depending on IS_CLOUD).
    """
    data = blob.get("data", blob)
    user = reconcile(data, structure)

    # Restore embedded attachments
    for sheet, items in user.get("checklists", {}).items():
        for idx, item in enumerate(items):
            new_atts = []
            for att in item.get("attachments", []):
                b64 = att.pop("data_b64", None)
                if b64 is not None:
                    raw = base64.b64decode(b64)
                    # Store via the normal path (respects IS_CLOUD and encryption)
                    stored_name = att["stored_name"]
                    if IS_CLOUD:
                        # Store plaintext in memory (encryption is local-only)
                        st.session_state.setdefault("_attach_bytes", {})[
                            _attach_mem_key(mode, sheet, stored_name)
                        ] = raw
                    else:
                        # Write to disk — re-encrypt if a fernet is active
                        f = get_fernet(mode)
                        d = item_attach_dir(mode, sheet)
                        if f is not None:
                            if not stored_name.endswith(".enc"):
                                stored_name = stored_name + ".enc"
                            (d / stored_name).write_bytes(f.encrypt(raw))
                            att["stored_name"] = stored_name
                            att["encrypted"] = True
                        else:
                            if stored_name.endswith(".enc"):
                                stored_name = stored_name[:-4]
                            (d / stored_name).write_bytes(raw)
                            att["stored_name"] = stored_name
                            att["encrypted"] = False
                new_atts.append(att)
            item["attachments"] = new_atts

    st.session_state["user_data"][mode] = user
    if "someone" in mode.lower() and "someone_name" in blob:
        st.session_state["someone_name"] = blob["someone_name"]


def reencrypt_all_attachments(mode: str, old_fernet: Optional[Fernet], new_fernet: Optional[Fernet]):
    """Re-encrypt / decrypt / rotate all attachments in place."""
    if IS_CLOUD:
        mem = st.session_state.get("_attach_bytes", {})
        prefix = f"{mode_slug(mode)}:"
        for key in list(mem.keys()):
            if not key.startswith(prefix):
                continue
            try:
                raw = mem[key]
                was_encrypted = key.endswith(".enc")
                if was_encrypted and old_fernet is not None:
                    raw = old_fernet.decrypt(raw)
                elif was_encrypted and old_fernet is None:
                    continue
                if new_fernet is not None:
                    new_data = new_fernet.encrypt(raw)
                    new_key = key if was_encrypted else key + ".enc"
                else:
                    new_data = raw
                    new_key = key[:-4] if was_encrypted else key
                if new_key != key:
                    del mem[key]
                mem[new_key] = new_data
            except Exception:
                continue
    else:
        root = ATTACH_DIR / mode_slug(mode)
        if not root.exists():
            return
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            try:
                data = p.read_bytes()
                was_encrypted = p.name.endswith(".enc")
                if was_encrypted and old_fernet is not None:
                    data = old_fernet.decrypt(data)
                elif was_encrypted and old_fernet is None:
                    continue
                if new_fernet is not None:
                    new_data = new_fernet.encrypt(data)
                    new_name = p.name if was_encrypted else p.name + ".enc"
                else:
                    new_data = data
                    new_name = p.name[:-4] if was_encrypted else p.name
                new_path = p.parent / new_name
                new_path.write_bytes(new_data)
                if new_path != p:
                    p.unlink()
            except Exception:
                continue

    # Update JSON metadata
    data = st.session_state["user_data"][mode]
    for sheet, items in data["checklists"].items():
        for it in items:
            for att in it.get("attachments", []):
                stored = att["stored_name"]
                was_encrypted = stored.endswith(".enc")
                if new_fernet is not None:
                    att["encrypted"] = True
                    att["stored_name"] = stored if was_encrypted else stored + ".enc"
                else:
                    att["encrypted"] = False
                    att["stored_name"] = stored[:-4] if was_encrypted else stored


# ---------------- PDF Export ----------------

def esc(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_pdf(structure: dict, user: dict, mode: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        title="Death Planning Workbook",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=colors.HexColor(PRIMARY))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor(PRIMARY))
    body = styles["BodyText"]
    small = ParagraphStyle("small", parent=body, fontSize=8, textColor=colors.HexColor(MUTED))

    story = []
    story.append(Paragraph("Death Planning Workbook", h1))
    story.append(Paragraph(f"Mode: {mode}", small))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", small))
    story.append(Spacer(1, 0.2 * inch))

    for sheet in CHECKLIST_SHEETS:
        items = structure["checklists"].get(sheet, [])
        ustate = user["checklists"].get(sheet, [])
        if not items:
            continue
        story.append(PageBreak())
        story.append(Paragraph(sheet, h1))
        by_cat: dict[str, list] = {}
        for i, it in enumerate(items):
            by_cat.setdefault(it["category"] or "Other", []).append((i, it))
        for cat, rows in by_cat.items():
            story.append(Spacer(1, 0.1 * inch))
            story.append(Paragraph(cat, h2))
            for i, it in rows:
                u = ustate[i] if i < len(ustate) else {}
                story.append(Paragraph(
                    f"<b>{esc(it['item'])}</b> — <i>{esc(u.get('status',''))}</i>", body,
                ))
                if it.get("details"):
                    story.append(Paragraph(esc(it["details"]), small))
                if it.get("notes"):
                    story.append(Paragraph("Ref: " + esc(it["notes"]), small))
                if u.get("user_info"):
                    story.append(Paragraph("Info: " + esc(u["user_info"]), body))
                if u.get("user_notes"):
                    story.append(Paragraph("Notes: " + esc(u["user_notes"]), body))
                atts = u.get("attachments", [])
                if atts:
                    names = ", ".join(esc(a["name"]) for a in atts)
                    story.append(Paragraph(f"Attached: {names}", small))
                story.append(Spacer(1, 0.05 * inch))

    for sheet in TABLE_SHEETS:
        tbl = structure["tables"].get(sheet)
        if not tbl:
            continue
        rows = user["tables"].get(sheet, [])
        story.append(PageBreak())
        story.append(Paragraph(sheet, h1))
        headers = tbl["headers"]
        data = [headers]
        for r in rows:
            data.append([esc(str(r.get(h, ""))) for h in headers])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(MUTED)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)

    doc.build(story)
    return buf.getvalue()


# ---------------- Form PDFs ----------------

def _form_doc(buf, title: str):
    return SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.9 * inch, rightMargin=0.9 * inch,
        topMargin=0.9 * inch, bottomMargin=0.9 * inch,
        title=title,
    )


def _form_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "title", parent=styles["Heading1"], fontSize=14, alignment=1,
            textColor=colors.black, spaceAfter=6,
        ),
        "subtitle": ParagraphStyle(
            "subtitle", parent=styles["Normal"], fontSize=9, alignment=1,
            textColor=colors.HexColor(MUTED), spaceAfter=14,
        ),
        "section": ParagraphStyle(
            "section", parent=styles["Heading3"], fontSize=10,
            textColor=colors.black, spaceBefore=10, spaceAfter=4,
        ),
        "body": ParagraphStyle(
            "body", parent=styles["BodyText"], fontSize=10,
            textColor=colors.black, leading=14,
        ),
        "small": ParagraphStyle(
            "small", parent=styles["BodyText"], fontSize=8,
            textColor=colors.HexColor(MUTED), leading=11,
        ),
        "sig": ParagraphStyle(
            "sig", parent=styles["BodyText"], fontSize=10,
            textColor=colors.black, spaceBefore=20,
        ),
    }


def _field_or_blank(val: str) -> str:
    s = (val or "").strip()
    return esc(s) if s else "_" * 60


def build_hipaa_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _form_doc(buf, "HIPAA Authorization")
    s = _form_styles()
    story = []
    story.append(Paragraph("HIPAA AUTHORIZATION FOR RELEASE OF PROTECTED HEALTH INFORMATION", s["title"]))
    story.append(Paragraph("Pursuant to 45 CFR § 164.508", s["subtitle"]))

    story.append(Paragraph("<b>1. Patient Information</b>", s["section"]))
    story.append(Paragraph(f"Name: {_field_or_blank(data.get('patient_name'))}", s["body"]))
    story.append(Paragraph(f"Date of birth: {_field_or_blank(data.get('patient_dob'))}", s["body"]))
    story.append(Paragraph(f"Phone: {_field_or_blank(data.get('patient_phone'))}", s["body"]))
    story.append(Paragraph(f"Address: {_field_or_blank(data.get('patient_address'))}", s["body"]))

    story.append(Paragraph("<b>2. Provider Authorized to Release Records</b>", s["section"]))
    story.append(Paragraph(_field_or_blank(data.get('provider_name')), s["body"]))

    story.append(Paragraph("<b>3. Person or Entity Authorized to Receive Records</b>", s["section"]))
    story.append(Paragraph(_field_or_blank(data.get('recipient_name')), s["body"]))

    story.append(Paragraph("<b>4. Information to be Released</b>", s["section"]))
    story.append(Paragraph(_field_or_blank(data.get('info_description')), s["body"]))

    story.append(Paragraph("<b>5. Purpose of Disclosure</b>", s["section"]))
    story.append(Paragraph(_field_or_blank(data.get('purpose')), s["body"]))

    story.append(Paragraph("<b>6. Special Categories of Information</b>", s["section"]))
    story.append(Paragraph(
        "Some categories of information receive heightened protection under federal or state law. "
        "Check any that may be released under this authorization:",
        s["small"],
    ))
    def _box(v): return "[X]" if v else "[ ]"
    story.append(Paragraph(
        f"{_box(data.get('inc_psych'))} Mental health and psychotherapy notes (may require separate authorization)", s["body"],
    ))
    story.append(Paragraph(
        f"{_box(data.get('inc_hiv'))} HIV / AIDS-related information", s["body"],
    ))
    story.append(Paragraph(
        f"{_box(data.get('inc_substance'))} Alcohol and substance use treatment records (42 CFR Part 2)", s["body"],
    ))
    story.append(Paragraph(
        f"{_box(data.get('inc_genetic'))} Genetic information", s["body"],
    ))

    story.append(Paragraph("<b>7. Expiration</b>", s["section"]))
    story.append(Paragraph(
        f"This authorization expires: {_field_or_blank(data.get('expiration'))}",
        s["body"],
    ))

    story.append(Paragraph("<b>8. Right to Revoke</b>", s["section"]))
    story.append(Paragraph(
        "I understand that I may revoke this authorization in writing at any time by sending a written "
        "notice to the provider named in Section 2, except to the extent that action has already been "
        "taken in reliance on it.",
        s["body"],
    ))

    story.append(Paragraph("<b>9. Redisclosure</b>", s["section"]))
    story.append(Paragraph(
        "I understand that information disclosed pursuant to this authorization may be redisclosed by "
        "the recipient and no longer protected by federal privacy regulations.",
        s["body"],
    ))

    story.append(Paragraph("<b>10. Conditions on Treatment</b>", s["section"]))
    story.append(Paragraph(
        "I understand that my healthcare provider may not condition treatment, payment, enrollment in "
        "a health plan, or eligibility for benefits on whether I sign this authorization, except as "
        "permitted by law.",
        s["body"],
    ))

    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph("<b>Signature</b>", s["section"]))

    is_rep = bool(data.get("signer_is_representative"))
    sig_table = Table(
        [
            ["Signature:", "_" * 40, "Date:", "_" * 20],
        ],
        colWidths=[0.9 * inch, 2.7 * inch, 0.6 * inch, 1.6 * inch],
    )
    sig_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(sig_table)

    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        f"Printed name: {_field_or_blank(data.get('patient_name'))}",
        s["body"],
    ))
    if is_rep:
        story.append(Spacer(1, 0.15 * inch))
        story.append(Paragraph(
            "Signed by personal representative on behalf of the patient.",
            s["small"],
        ))
        story.append(Paragraph(
            f"Representative name: {_field_or_blank(data.get('representative_name'))}",
            s["body"],
        ))
        story.append(Paragraph(
            f"Relationship / legal authority: {_field_or_blank(data.get('representative_relationship'))}",
            s["body"],
        ))

    doc.build(story)
    return buf.getvalue()


def build_letter_of_instruction_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _form_doc(buf, "Letter of Instruction")
    s = _form_styles()
    story = []
    story.append(Paragraph("Letter of Instruction", s["title"]))
    story.append(Paragraph(
        f"{esc(data.get('writer_name') or '')}  ·  {esc(data.get('date') or '')}",
        s["subtitle"],
    ))

    def sec(label: str, key: str):
        val = (data.get(key) or "").strip()
        if not val:
            return
        story.append(Paragraph(f"<b>{esc(label)}</b>", s["section"]))
        for para in val.split("\n\n"):
            story.append(Paragraph(esc(para).replace("\n", "<br/>"), s["body"]))

    story.append(Paragraph(esc(data.get('addressed_to') or 'To my family and executor'), s["body"]))
    story.append(Spacer(1, 0.1 * inch))
    if data.get("opening"):
        story.append(Paragraph(esc(data['opening']).replace("\n", "<br/>"), s["body"]))

    sec("Where the will and trust documents are", "location_of_will")
    sec("Where financial records are", "location_of_financial")
    sec("Passwords and the password manager", "location_of_passwords")
    sec("Who to call first", "first_calls")
    sec("Summary of accounts and assets", "financial_summary")
    sec("Debts and ongoing bills", "debts")
    sec("Digital accounts and instructions", "digital_assets")
    sec("Funeral and burial wishes", "funeral_wishes")
    sec("Personal messages", "personal_messages")

    if data.get("closing"):
        story.append(Spacer(1, 0.2 * inch))
        story.append(Paragraph(esc(data['closing']).replace("\n", "<br/>"), s["body"]))
        story.append(Spacer(1, 0.2 * inch))
        story.append(Paragraph(esc(data.get('writer_name') or ''), s["body"]))

    doc.build(story)
    return buf.getvalue()


def build_funeral_wishes_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _form_doc(buf, "Funeral and Burial Wishes")
    s = _form_styles()
    story = []
    story.append(Paragraph("Funeral and Burial Wishes", s["title"]))
    story.append(Paragraph(
        f"{esc(data.get('name') or '')}  ·  {esc(data.get('date') or '')}",
        s["subtitle"],
    ))
    story.append(Paragraph(
        "This is a statement of my preferences. It is not legally binding but reflects my wishes. "
        "I ask that my family honor them to the extent practical.",
        s["small"],
    ))

    fields = [
        ("Disposition", "disposition"),
        ("Location", "location"),
        ("Pre-paid arrangements", "prepaid"),
        ("Service type", "service_type"),
        ("Officiant", "officiant"),
        ("Service location", "location_of_service"),
        ("Music", "music"),
        ("Readings, speakers, or prayers", "readings"),
        ("Flowers or charitable donations", "flowers_donations"),
        ("Obituary", "obituary"),
        ("People and organizations to notify", "notify"),
        ("Additional notes", "additional"),
    ]
    for label, key in fields:
        val = (data.get(key) or "").strip()
        if not val:
            continue
        story.append(Paragraph(f"<b>{esc(label)}</b>", s["section"]))
        story.append(Paragraph(esc(val).replace("\n", "<br/>"), s["body"]))

    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph("Signature: " + "_" * 40 + "     Date: " + "_" * 20, s["body"]))

    doc.build(story)
    return buf.getvalue()


def build_digital_assets_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = _form_doc(buf, "Digital Asset Instructions")
    s = _form_styles()
    story = []
    story.append(Paragraph("Digital Asset Instructions", s["title"]))
    story.append(Paragraph(
        f"{esc(data.get('writer_name') or '')}  ·  {esc(data.get('date') or '')}",
        s["subtitle"],
    ))
    story.append(Paragraph(
        "These are my instructions for handling my online accounts and digital property. "
        "This document is not a substitute for provisions in my will, but is intended to "
        "guide my executor and fiduciary.",
        s["small"],
    ))

    fields = [
        ("Password manager", "password_manager"),
        ("Primary email", "primary_email"),
        ("Phone and 2FA", "phone_accounts"),
        ("Financial account logins", "financial_logins"),
        ("Social media", "social_media"),
        ("Cloud storage and photos", "cloud_storage"),
        ("Subscriptions", "subscriptions"),
        ("Cryptocurrency and wallets", "crypto"),
        ("Legacy contacts", "legacy_contacts"),
        ("Other notes", "notes"),
    ]
    for label, key in fields:
        val = (data.get(key) or "").strip()
        if not val:
            continue
        story.append(Paragraph(f"<b>{esc(label)}</b>", s["section"]))
        story.append(Paragraph(esc(val).replace("\n", "<br/>"), s["body"]))

    doc.build(story)
    return buf.getvalue()


FORM_BUILDERS = {
    "hipaa": build_hipaa_pdf,
    "letter_of_instruction": build_letter_of_instruction_pdf,
    "funeral_wishes": build_funeral_wishes_pdf,
    "digital_assets": build_digital_assets_pdf,
}


# ---------------- UI helpers ----------------

def section_progress(items, ustate):
    if not items:
        return 0.0, 0, 0
    counted = [u for u in ustate if u.get("status") != "N/A"]
    done = sum(1 for u in counted if u.get("status") == "Done")
    total = len(counted) or 1
    return done / total, done, total


SECTION_GUIDANCE_KEYS = {
    "Before": "before_urgency",
    "During Illness": "during_intro",
    "At End - Hospice": "end_intro",
    "After Death": "after_intro",
}


def render_checklist(sheet: str, structure: dict):
    c = ctx()
    guidance_key = SECTION_GUIDANCE_KEYS.get(sheet)
    if guidance_key and guidance_key in c:
        st.markdown(f"*{c[guidance_key]}*")

    items = structure["checklists"].get(sheet, [])
    mode = mode_key()
    ustate = st.session_state["user_data"][mode]["checklists"].setdefault(
        sheet, [{"status": "Not started", "user_notes": "", "user_info": "", "attachments": []} for _ in items]
    )
    while len(ustate) < len(items):
        ustate.append({"status": "Not started", "user_notes": "", "user_info": "", "attachments": []})

    pct, done, total = section_progress(items, ustate)
    st.progress(pct, text=f"{done} of {total} complete")

    by_cat: dict[str, list] = {}
    for i, it in enumerate(items):
        key = it.get("timeframe") or it["category"] or "Other"
        by_cat.setdefault(key, []).append((i, it))

    for cat, rows in by_cat.items():
        st.markdown(f"### {cat}")
        for i, it in rows:
            u = ustate[i]
            u.setdefault("attachments", [])
            n_att = len(u["attachments"])
            att_tag = f"  📎{n_att}" if n_att else ""
            with st.expander(f"**{it['item']}**  ·  _{u['status']}_{att_tag}", expanded=False):
                if it.get("details"):
                    st.markdown(f"_{it['details']}_")
                if it.get("notes"):
                    st.caption(f"Reference: {it['notes']}")

                c1, c2 = st.columns([1, 3])
                with c1:
                    u["status"] = st.selectbox(
                        "Status", STATUSES,
                        index=STATUSES.index(u["status"]) if u["status"] in STATUSES else 0,
                        key=f"{mode}:{sheet}:{i}:status",
                    )
                with c2:
                    u["user_info"] = st.text_input(
                        "Your info (location, reference — e.g., 'in 1Password', 'safe deposit box')",
                        value=u.get("user_info", ""),
                        key=f"{mode}:{sheet}:{i}:info",
                    )
                u["user_notes"] = st.text_area(
                    "Notes",
                    value=u.get("user_notes", ""),
                    key=f"{mode}:{sheet}:{i}:notes",
                    height=70,
                )

                # Attachments
                st.markdown("**Attached documents**")
                if u["attachments"]:
                    for ai, att in enumerate(list(u["attachments"])):
                        ac1, ac2, ac3 = st.columns([5, 2, 1])
                        with ac1:
                            size_kb = max(1, att.get("size", 0) // 1024)
                            lock = " 🔒" if att.get("encrypted") else ""
                            st.caption(f"📎 {att['name']} · {size_kb} KB{lock}")
                        with ac2:
                            blob = read_attachment(mode, sheet, att)
                            if blob is not None:
                                st.download_button(
                                    "Download", data=blob, file_name=att["name"],
                                    key=f"dl:{mode}:{sheet}:{i}:{ai}",
                                    use_container_width=True,
                                )
                            else:
                                st.caption("(locked)")
                        with ac3:
                            if st.button("✕", key=f"rm:{mode}:{sheet}:{i}:{ai}",
                                         help="Remove attachment"):
                                delete_attachment(mode, sheet, att)
                                u["attachments"].pop(ai)
                                save_user_data(mode)
                                st.rerun()
                else:
                    st.caption("_No attachments yet._")

                with st.form(key=f"upform:{mode}:{sheet}:{i}", clear_on_submit=True):
                    new_file = st.file_uploader(
                        "Attach a document (PDF, image, etc.)",
                        key=f"up:{mode}:{sheet}:{i}",
                        label_visibility="collapsed",
                    )
                    if st.form_submit_button("Attach file"):
                        if new_file is not None:
                            rec = save_attachment(mode, sheet, i, new_file)
                            u["attachments"].append(rec)
                            save_user_data(mode)
                            st.rerun()
                        else:
                            st.warning("Pick a file first.")


def render_table(sheet: str, structure: dict):
    c = ctx()
    if sheet == "Key Contacts":
        st.markdown(f"*{c['contacts_intro']}*")
    elif sheet == "Account Registry":
        st.markdown(f"*{c['registry_intro']}*")

    tbl = structure["tables"].get(sheet)
    if not tbl:
        st.info("No table defined.")
        return
    mode = mode_key()
    rows = st.session_state["user_data"][mode]["tables"].setdefault(
        sheet, [dict(r) for r in tbl["seed"]]
    )
    headers = tbl["headers"]

    st.caption("Add, edit, or remove rows. Click **Save** in the sidebar to persist changes.")
    import pandas as pd
    df = pd.DataFrame(rows)
    for h in headers:
        if h not in df.columns:
            df[h] = ""
    df = df[headers + [c for c in df.columns if c not in headers and c != "_section"]]
    edited = st.data_editor(
        df, num_rows="dynamic", use_container_width=True,
        key=f"{mode}:{sheet}:editor",
    )
    st.session_state["user_data"][mode]["tables"][sheet] = edited.fillna("").to_dict("records")


def render_forms(structure: dict):
    mode = mode_key()
    user = st.session_state["user_data"][mode]
    user.setdefault("forms", {})

    st.markdown("## Forms")
    st.markdown(
        "Fillable templates for the documents that *can* be standardized. "
        "Fill them in here, then generate a PDF to print and sign."
    )
    st.warning(
        "**Important.** Of the documents in this space, only the HIPAA authorization "
        "is genuinely nationwide — it's defined by federal rule (45 CFR 164.508). "
        "Advance directives, living wills, healthcare and financial powers of attorney, "
        "and DNR/POLST orders are **state-specific** and this app does not generate them. "
        "Use your state's official form for those. See the bottom of this page for pointers. "
        "If your situation is complicated, talk to an estate attorney — this app is not a substitute."
    )

    form_ids = list(FORMS.keys())
    labels = [FORMS[fid]["title"] for fid in form_ids]
    choice_idx = st.selectbox(
        "Choose a form",
        range(len(form_ids)),
        format_func=lambda i: labels[i],
        key=f"form_choice_{mode}",
    )
    form_id = form_ids[choice_idx]
    spec = FORMS[form_id]

    c = ctx()
    st.markdown(f"### {spec['title']}")
    st.caption(spec["blurb"])
    if spec["kind"] == "legal-federal":
        st.info(
            "This form is governed by federal regulation (45 CFR 164.508) and is valid in all states "
            "when completed and signed. Some providers may still require their own form — that's fine, "
            "this one works as a backup or starting point."
        )
    else:
        st.info(
            f"This is a **personal document**, not a legal instrument. It's meant to supplement "
            f"{c['your_or_their']} will and help the people acting on {c['your_or_their']} behalf. "
            f"It does not replace a will or any state-required legal form."
        )

    # Mode-specific guidance per form
    form_ctx_key = {
        "hipaa": "forms_hipaa",
        "letter_of_instruction": "forms_letter",
        "funeral_wishes": "forms_funeral",
        "digital_assets": "forms_digital",
    }.get(form_id)
    if form_ctx_key and form_ctx_key in c:
        st.markdown(f"*{c[form_ctx_key]}*")

    form_data = user["forms"].setdefault(
        form_id, {f[0]: f[3] for f in spec["fields"]}
    )
    # Ensure all current fields exist (in case spec changed)
    for f in spec["fields"]:
        form_data.setdefault(f[0], f[3])

    st.divider()
    for field_key, label, field_type, default in spec["fields"]:
        widget_key = f"form:{mode}:{form_id}:{field_key}"
        if field_type == "text":
            form_data[field_key] = st.text_input(
                label, value=form_data.get(field_key, default) or "", key=widget_key,
            )
        elif field_type == "textarea":
            form_data[field_key] = st.text_area(
                label, value=form_data.get(field_key, default) or "",
                key=widget_key, height=90,
            )
        elif field_type == "checkbox":
            form_data[field_key] = st.checkbox(
                label, value=bool(form_data.get(field_key, default)), key=widget_key,
            )

    st.divider()
    col_a, col_b = st.columns([1, 1])
    with col_a:
        if st.button("📄 Generate PDF", key=f"form_pdf_{form_id}", use_container_width=True):
            try:
                pdf_bytes = FORM_BUILDERS[form_id](form_data)
                st.session_state[f"_form_pdf_{form_id}"] = pdf_bytes
                save_user_data(mode)
                st.success("PDF ready. Click the download button.")
            except Exception as e:
                st.error(f"Couldn't build PDF: {e}")
    with col_b:
        pdf_bytes = st.session_state.get(f"_form_pdf_{form_id}")
        if pdf_bytes:
            st.download_button(
                "⬇ Download PDF",
                data=pdf_bytes,
                file_name=f"{form_id}_{mode_slug(mode)}.pdf",
                mime="application/pdf",
                key=f"form_dl_{form_id}",
                use_container_width=True,
            )

    st.divider()
    with st.expander("State-specific forms — where to find them"):
        st.caption(
            "These documents must come from your state. Generating them from a template "
            "risks creating something that won't be honored when it matters."
        )
        for name, desc, where in STATE_FORM_POINTERS:
            st.markdown(f"**{name}** — {desc}  \n_{where}_")


def render_report(structure: dict):
    c = ctx()
    mode = mode_key()
    user = st.session_state["user_data"][mode]
    st.markdown("## Completeness report")
    st.caption(
        f"What's still open. **Before** items are flagged most aggressively — "
        f"{c['report_before_flag']}"
    )

    overall_done = overall_total = 0
    for sheet in CHECKLIST_SHEETS:
        items = structure["checklists"].get(sheet, [])
        ustate = user["checklists"].get(sheet, [])
        pct, d, t = section_progress(items, ustate)
        overall_done += d
        overall_total += t
        st.markdown(f"**{sheet}** — {d}/{t}")
        st.progress(pct)

    st.divider()
    st.markdown(f"### Overall: {overall_done} / {overall_total}")
    st.progress((overall_done / overall_total) if overall_total else 0)

    st.divider()
    st.markdown("### Missing items by priority")
    priority_order = ["Before", "During Illness", "At End - Hospice", "After Death"]
    for sheet in priority_order:
        items = structure["checklists"].get(sheet, [])
        ustate = user["checklists"].get(sheet, [])
        missing = [
            (i, it) for i, it in enumerate(items)
            if i < len(ustate) and ustate[i]["status"] in ("Not started", "In progress")
        ]
        if not missing:
            continue
        flag = "🔴" if sheet == "Before" else ("🟠" if sheet == "During Illness" else "🟡")
        with st.expander(
            f"{flag} {sheet} — {len(missing)} open", expanded=(sheet == "Before"),
        ):
            if sheet == "Before":
                st.caption(c["report_before_flag"])
            for i, it in missing:
                status = ustate[i]["status"]
                st.markdown(f"- **{it['item']}** _({status})_ — {it.get('details','')}")


def render_security_sidebar(structure: dict):
    mode = mode_key()
    st.markdown("### Security")
    current = get_fernet(mode)
    encrypted_on_disk = file_is_encrypted(data_file(mode))

    if current is not None:
        st.success("🔒 This workbook is encrypted.")
        if st.button("Lock (clear passphrase from memory)", use_container_width=True, key=f"lock_{mode}"):
            set_fernet(mode, None)
            st.session_state["user_data"][mode] = default_user_state(structure)
            st.rerun()
        with st.expander("Change or remove passphrase"):
            st.caption("Rotate to a new passphrase or remove encryption entirely.")
            new1 = st.text_input("New passphrase", type="password", key=f"rot1_{mode}")
            new2 = st.text_input("Confirm", type="password", key=f"rot2_{mode}")
            if st.button("Rotate passphrase", key=f"rot_btn_{mode}"):
                if not new1 or new1 != new2:
                    st.error("Passphrases don't match.")
                elif len(new1) < 8:
                    st.error("Use at least 8 characters.")
                else:
                    old = current
                    salt = os.urandom(16)
                    new_key = derive_key(new1, salt)
                    new_f = Fernet(new_key)
                    reencrypt_all_attachments(mode, old, new_f)
                    set_fernet(mode, new_f)
                    set_salt(mode, base64.b64encode(salt).decode())
                    save_user_data(mode)
                    st.success("Passphrase rotated.")
                    st.rerun()
            if st.button("Remove encryption (save as plaintext)", key=f"dec_{mode}"):
                reencrypt_all_attachments(mode, current, None)
                set_fernet(mode, None)
                set_salt(mode, None)
                save_user_data(mode)
                st.warning("Saved as plaintext.")
                st.rerun()
    elif encrypted_on_disk:
        st.warning("🔒 Saved file is encrypted. Enter passphrase to unlock.")
        pp = st.text_input("Passphrase", type="password", key=f"unl_{mode}")
        if st.button("Unlock", key=f"unl_btn_{mode}", use_container_width=True):
            data, status = load_user_data(mode, structure, passphrase=pp)
            if status == "loaded":
                st.session_state["user_data"][mode] = data
                st.success("Unlocked.")
                st.rerun()
            elif status == "bad_passphrase":
                st.error("Wrong passphrase.")
    else:
        st.caption("🔓 Not encrypted. The saved file is plaintext JSON.")
        with st.expander("Set a passphrase"):
            st.caption(
                "This encrypts the saved file and any uploaded documents. "
                "If you forget the passphrase, the data is **unrecoverable**."
            )
            p1 = st.text_input("New passphrase", type="password", key=f"set1_{mode}")
            p2 = st.text_input("Confirm", type="password", key=f"set2_{mode}")
            if st.button("Enable encryption", key=f"set_btn_{mode}"):
                if not p1 or p1 != p2:
                    st.error("Passphrases don't match.")
                elif len(p1) < 8:
                    st.error("Use at least 8 characters.")
                else:
                    salt = os.urandom(16)
                    key = derive_key(p1, salt)
                    f = Fernet(key)
                    reencrypt_all_attachments(mode, None, f)
                    set_fernet(mode, f)
                    set_salt(mode, base64.b64encode(salt).decode())
                    save_user_data(mode)
                    st.success("Encryption enabled. Don't lose the passphrase.")
                    st.rerun()


# ---------------- Main ----------------

def main():
    st.set_page_config(
        page_title="Death Planning Workbook",
        page_icon="🕊",
        layout="wide",
    )
    st.markdown(
        f"""
        <style>
        .stApp {{ background-color: {BG}; color: #1F2A36; }}
        .stApp p, .stApp li, .stApp label, .stApp span, .stApp div {{ color: #1F2A36; }}
        h1, h2, h3, h4 {{ color: {PRIMARY} !important; }}
        .stProgress > div > div > div > div {{ background-color: {PRIMARY}; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    if not XLSX_PATH.exists():
        st.error(f"Missing workbook file: {XLSX_PATH.name}")
        return

    structure = load_workbook_structure()

    # Init state
    if "user_data" not in st.session_state:
        st.session_state["user_data"] = {}
        for m in MODES:
            data, status = load_user_data(m, structure, passphrase=None)
            st.session_state["user_data"][m] = data
    if "mode" not in st.session_state:
        st.session_state["mode"] = "For my parents"

    # Header
    st.title("Death Planning Workbook")
    st.caption(
        "A practical binder for the hard stuff. Work through it at your own pace — "
        "nothing here needs to be done all at once. Save often."
    )

    col_mode, col_name = st.columns([3, 4])
    with col_mode:
        st.session_state["mode"] = st.radio(
            "This workbook is",
            MODES,
            horizontal=True,
            index=MODES.index(st.session_state.get("mode", MODES[0])),
        )
    with col_name:
        if st.session_state["mode"] == "For someone else":
            st.session_state["someone_name"] = st.text_input(
                "Their name (used in guidance text)",
                value=st.session_state.get("someone_name", ""),
                placeholder="e.g., Aunt Linda, my brother David",
            )
    mode = mode_key()
    locked = (not IS_CLOUD
              and file_is_encrypted(data_file(mode))
              and get_fernet(mode) is None)

    # Cloud banner
    if IS_CLOUD:
        st.info(
            "☁️ **You're using the online preview.** Your data lives in this browser session only — "
            "**download the JSON** before closing the tab or it's gone. "
            "For long-term use, run the app locally (see the GitHub repo)."
        )

    # Sidebar
    with st.sidebar:
        st.markdown("### Sections")
        nav_options = (
            ["Overview"] + CHECKLIST_SHEETS + TABLE_SHEETS
            + ["Forms", "Completeness report"]
        )
        choice = st.radio("Navigate", nav_options, label_visibility="collapsed")

        st.divider()
        if not IS_CLOUD:
            render_security_sidebar(structure)
            st.divider()

        # -- Save & export --
        if IS_CLOUD:
            st.markdown("### Save your work")
            st.caption(
                "Your data lives in this browser tab only. "
                "**Download the JSON** to keep it. Upload it next time to continue."
            )
        else:
            st.markdown("### Save & export")
            # Auto-save on every interaction (local only)
            if not locked:
                save_user_data(mode)
                saved_at = datetime.now().strftime("%H:%M:%S")
                st.caption(f"Auto-saved at {saved_at}")

        dl_obj = export_portable_json(mode) if not locked else {}
        payload = json.dumps(dl_obj, indent=2) if dl_obj else "{}"
        payload_mb = len(payload) / (1024 * 1024)
        size_note = f" ({payload_mb:.1f} MB)" if payload_mb > 0.5 else ""
        if IS_CLOUD:
            st.download_button(
                f"⬇ Download JSON — your only save{size_note}",
                data=payload,
                file_name=f"death_workbook_{mode_slug(mode)}.json",
                mime="application/json",
                use_container_width=True,
                disabled=locked,
                type="primary",
            )
            st.caption("Includes all attached documents.")
        else:
            st.download_button(
                f"⬇ Download JSON backup{size_note}",
                data=payload,
                file_name=f"death_workbook_{mode_slug(mode)}.json",
                mime="application/json",
                use_container_width=True,
                disabled=locked,
                help="Complete backup including attachments. Use to share or move to another machine.",
            )
        uploaded = st.file_uploader("Load JSON", type=["json"])
        if uploaded is not None and not locked:
            try:
                blob = json.loads(uploaded.read())
                import_portable_json(blob, mode, structure)
                st.success("Loaded (including any attachments).")
            except Exception as e:
                st.error(f"Couldn't load: {e}")

        st.divider()
        if st.button("📄 Generate PDF", use_container_width=True, disabled=locked):
            pdf_bytes = build_pdf(structure, st.session_state["user_data"][mode], mode)
            st.session_state["_pdf_bytes"] = pdf_bytes
        if st.session_state.get("_pdf_bytes"):
            st.download_button(
                "⬇ Download PDF",
                data=st.session_state["_pdf_bytes"],
                file_name=f"death_workbook_{mode_slug(mode)}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        st.divider()
        st.caption(
            "Never put raw passwords in this file. Write *'in 1Password, item X'* or "
            "*'safe deposit box'* instead."
        )

    # Main panel
    if locked:
        st.warning(
            f"🔒 **{mode}** is encrypted. Enter the passphrase in the sidebar under "
            "**Security** to unlock. The workbook is empty until unlocked."
        )
        return

    if choice == "Overview":
        render_overview(structure)
    elif choice in CHECKLIST_SHEETS:
        st.markdown(f"## {choice}")
        render_checklist(choice, structure)
    elif choice in TABLE_SHEETS:
        st.markdown(f"## {choice}")
        render_table(choice, structure)
    elif choice == "Forms":
        render_forms(structure)
    elif choice == "Completeness report":
        render_report(structure)


def render_overview(structure: dict):
    c = ctx()
    st.markdown("## Where to start")
    st.markdown(c["overview_intro"])
    st.markdown(
        f"""
        The workbook is organized around the sequence things actually happen in:

        1. **Before** — what to gather *now*, while {c['subject']} {c['subject_they'] == 'you' and 'are' or 'is'} healthy. {c['before_urgency'].split('.')[0]}.
        2. **During Illness** — care management, insurance, finances during a decline.
        3. **At End / Hospice** — DNR, hospice enrollment, final wishes.
        4. **After Death** — organized by timeframe (first 48 hours → 6 months).
        5. **Key Contacts** — the people who'll need to be called.
        6. **Account Registry** — every account, policy, and subscription in one place.
        7. **Forms** — fillable templates (HIPAA authorization, letter of instruction, funeral wishes, digital asset instructions) to print and sign.

        **A few things worth knowing:**
        - Work one item at a time. Nothing here needs to happen in one sitting.
        - Use the **"Your info"** box next to each item to write where something lives.
        - You can **attach scanned documents** (will, POA, deed, etc.) to any item.
        - Save often. {'Download the JSON before closing this tab.' if IS_CLOUD else 'The **Save** button in the sidebar writes to a local JSON file.'}
        - The **Completeness report** tells you what's still open, with the "Before" items flagged most urgently.
        - When ready, **Generate PDF** makes a printable binder.
        """
    )

    st.markdown("### A note on what to put in this file")
    st.markdown(
        f"""
        This file can hold sensitive information — SSNs, account numbers, medical details, scans of legal documents. Please read this before filling it in.

        **The best approach for really sensitive items** (SSNs, master passwords, financial account credentials):
        put them in a password manager like 1Password and then write a **reference** in this workbook
        ("SSN: in 1Password, item '{c['subject_possessive'].rstrip("'s") if c['subject_possessive'].endswith("'s") else c['subject']} SSN'").
        That way the workbook can be shared with family members or a lawyer later without exposing everything.
        """
    )
    if not IS_CLOUD:
        st.markdown(
            """
            **On encryption:** When you set a passphrase in the sidebar, the saved JSON file and any uploaded
            documents are encrypted on disk (AES, key derived from your passphrase). If you forget
            the passphrase, the data is gone — there is no recovery. Run this on a computer with full-disk
            encryption (FileVault / BitLocker) as the baseline.

            **On shared access:** if multiple family members need to work on this, keep the file in a
            shared encrypted vault (1Password, Bitwarden, Proton Drive) and sync manually.
            """
        )

    if structure["instructions"]:
        with st.expander("Workbook introduction (from the source file)"):
            for line in structure["instructions"]:
                st.markdown(line)


if __name__ == "__main__":
    main()
